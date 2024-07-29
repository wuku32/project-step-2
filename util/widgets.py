import re
import random
from enum import Enum
from typing import List
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook


class Status(Enum):
    UNUSED = 'Unused'  # 表示数据行从未被使用过
    IN_USE = 'InUse'  # 表示数据行当前正在被使用
    ARCHIVED = 'Archived'  # 表示数据行暂时无法使用，但保留以备历史查询
    DELETED = 'Deleted'  # 表示数据行已被标记为删除，但可能还没有从数据库中物理删除
    RECALLED = 'Recalled'  # 表示sku没有出单，召回
    OTHER = 'Other'  # 表示数据行已被已经挑选上，但是由于其他原因（如主图底色不是白色而弃选）最终放弃


def note_inuse_sku(df: DataFrame) -> DataFrame:
    # 已下架的sku
    inuse_skus = open('./util/skus/archived_skus.txt', mode='r', encoding='utf-8').readlines()
    inuse_skus = [i.strip() for i in inuse_skus]
    df.loc[df['productSku'].isin(inuse_skus), 'status'] = Status.IN_USE.value
    return df


def note_archived_sku(df: DataFrame) -> DataFrame:
    # 已下架的sku
    archived_skus = open('./util/skus/archived_skus.txt', mode='r', encoding='utf-8').readlines()
    archived_skus = [i.strip() for i in archived_skus]
    df.loc[df['productSku'].isin(archived_skus), 'status'] = Status.ARCHIVED.value
    return df


def note_deleted_sku(df: DataFrame) -> DataFrame:
    # 已下架的sku
    deleted_skus = open('./util/skus/deleted_skus.txt', mode='r', encoding='utf-8').readlines()
    deleted_skus = [i.strip() for i in deleted_skus]
    df.loc[df['productSku'].isin(deleted_skus), 'status'] = Status.DELETED.value
    return df


def note_recalled_sku(df: DataFrame) -> DataFrame:
    # 已下架的sku
    recalled_skus = open('./util/skus/other_skus.txt', mode='r', encoding='utf-8').readlines()
    recalled_skus = [i.strip() for i in recalled_skus]
    df.loc[df['productSku'].isin(recalled_skus), 'status'] = Status.RECALLED.value
    return df


def note_other_sku(df: DataFrame) -> DataFrame:
    # 已下架的sku
    other_skus = open('./util/skus/other_skus.txt', mode='r', encoding='utf-8').readlines()
    other_skus = [i.strip() for i in other_skus]
    df.loc[df['productSku'].isin(other_skus), 'status'] = Status.OTHER.value
    return df


def wash_rows(df: DataFrame) -> DataFrame:
    # 品牌名列表
    brands = open('./util/stopwords/brands.txt', mode='r', encoding='utf-8').readlines()
    # 非目标词列表
    minors = open('./util/stopwords/minors.txt', mode='r', encoding='utf-8').readlines()
    # 创建一个正则表达式，用于匹配品牌名或非目标词
    # 使用re.escape()确保特殊字符被正确转义
    keywords_pattern = '|'.join(re.escape(keyword.replace('\n', '')) for keyword in set(brands + minors))
    # 使用正则表达式来过滤包含关键词的行
    # 确保regex参数为True，这样正则表达式会被正确解析
    df = df[~df['term'].str.contains(keywords_pattern, case=False, regex=True)]
    return df


def remove_stopwords(text):
    # 确保传入的参数是字符串
    if not isinstance(text, str):
        text = str(text)

    # 品牌名列表
    brands = open('./util/stopwords/brands.txt', mode='r', encoding='utf-8').readlines()
    # 非目标词列表
    minors = open('./util/stopwords/minors.txt', mode='r', encoding='utf-8').readlines()

    # 创建一个正则表达式，用于匹配品牌名或非目标词
    # 使用re.escape()确保特殊字符被正确转义
    keywords_pattern = '|'.join(re.escape(keyword.replace('\n', '')) for keyword in set(brands + minors))
    regex = re.compile(r'\b(' + keywords_pattern + r')\b', re.IGNORECASE)
    return regex.sub('', text).strip()


def random_indexes(t: int, n: int = 20, x: int = 8, shuffle: bool = True) -> List[list]:
    k = int(t / x + 0.5)  # 每一堆的个数

    count = 0
    indexes = []

    while count != n:
        index = []
        for i in range(0, x):
            if i != x - 1:
                index.extend(random.sample(range(i * k, (i + 1) * k), 1))
                continue
            if (i * k) == t:
                continue
            index.extend(random.sample(range(i * k, t), 1))
        if shuffle:
            random.shuffle(index)
        indexes.append(index)
        count += 1
    return indexes


def generate_ym():
    from datetime import datetime, timedelta

    # 当前时间
    now = datetime.now()

    # 格式化日期
    year = now.year
    month = str(now.month).zfill(2)

    ym = f"{year}_{month}"
    return ym


def generate_ymd(delta: int = 0):
    from datetime import datetime, timedelta

    # 当前时间
    now = datetime.now()

    # 格式化日期
    year = now.year
    month = str(now.month).zfill(2)
    day = str((now + timedelta(days=delta)).day).zfill(2)

    ymd = f"{year}_{month}_{day}"
    return ymd


def generate_ymd_HMS():
    import time

    ymd_HMS = time.strftime('%Y-%m-%D %H:%M:%S', time.localtime())
    return ymd_HMS


def generate_i():
    from datetime import datetime, timedelta

    # 当前时间
    now = datetime.now()

    # 格式化起始小时
    start_hour = str(now.hour).zfill(2)

    # 计算结束小时
    end_hour = (now + timedelta(hours=1)).hour
    # end_hour = str(end_hour % 24).zfill(2)  # 确保24小时制转换成00
    end_hour = str(end_hour).zfill(2)

    i = f"({start_hour}-{end_hour})"
    return i

# def safe_write_in(fp: str, sheet_name: str, t: DataFrame, mode: str = 'insert'):
#     """
#     只更新工作表中指定的工作簿，而不影响其他工作簿原有的内容
#     :param fp: Excel文件路径
#     :param sheet_name: 工作簿名
#     :param t: 增/删/改的内容
#     :param mode: insert, update, delete
#     :return:
#     """
#     # 使用pandas读取Excel中的特定工作簿
#     try:
#         with pd.ExcelFile(fp) as xls:
#             df = xls.parse(sheet_name)
#     except FileNotFoundError:  # 第一次创建工作簿
#         df = pd.DataFrame()
#         df.to_excel(fp, sheet_name=sheet_name, index=False)
#     except:  # 第一次创建工作表
#         book = load_workbook(fp)
#         book.create_sheet(sheet_name)
#         book.save(fp)
#         with pd.ExcelFile(fp) as xls:
#             df = xls.parse(sheet_name)
#     book = load_workbook(fp)
#     # 修改DataFrame中的数据
#     if mode == 'insert':
#         # ignore_index=True会重新生成索引，这对于拼接操作来说通常是有用的，尤其是当两个DataFrame有重叠的索引时
#         df = pd.concat([df, t], axis=0, ignore_index=True)
#     if mode == 'update':
#         df = pd.DataFrame()
#         df = pd.concat([df, t.copy()], axis=0, ignore_index=True)
#         print(df)
#     if mode == 'delete':
#         df = df.merge(t, how='outer', indicator=True)
#         df = df[df['_merge'] == 'left_only']
#         df = df.drop(columns=['_merge'])
#     # 将修改后的DataFrame写回到Excel文件中的特定工作簿
#     writer = pd.ExcelWriter(fp, engine='openpyxl')
#     # writer = pd.ExcelWriter(fp, engine='xlsxwriter')
#     # 保持其他工作簿不变，只更新目标工作簿
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#     df.to_excel(writer, sheet_name=sheet_name, index=False)
#     writer.save()
